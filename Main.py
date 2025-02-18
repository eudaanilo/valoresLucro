import openpyxl

def calcular_valores():
    # Perguntar quantos itens
    num_itens = int(input("Quantos itens você deseja calcular? "))

    # Criar uma lista para armazenar os valores
    valores = []

    # Perguntar os valores dos itens
    for i in range(num_itens):
        valor = float(input(f"Digite o valor do item {i+1}: "))
        valores.append(valor)

    # Calcular os valores com os fatores
    fatores = [2.7, 2.5, 2.25]
    resultados = []

    for valor in valores:
        resultado = []
        for fator in fatores:
            resultado.append(valor * fator)
        resultados.append(resultado)

    # Imprimir a tabela
    print("\nTabela de Valores:")
    print("------------------------")
    print("Item\tValor Inserido\tValor x 2.7\tValor x 2.5\tValor x 2.25")
    print("------------------------")

    for i, (valor, resultado) in enumerate(zip(valores, resultados)):
        print(f"{i+1}\t{valor:.2f}\t\t{resultado[0]:.2f}\t\t{resultado[1]:.2f}\t\t{resultado[2]:.2f}")

    # Perguntar se o usuário deseja baixar a planilha
    resposta = input("\nDeseja baixar a planilha com a tabela de resultados finais? (y/n): ")

    if resposta.lower() == 'y':
        # Criar um arquivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active

        # Adicionar cabeçalho
        ws['A1'] = 'Item'
        ws['B1'] = 'Valor Inserido'
        ws['C1'] = 'Valor x 2.7'
        ws['D1'] = 'Valor x 2.5'
        ws['E1'] = 'Valor x 2.25'

        # Adicionar dados
        for i, (valor, resultado) in enumerate(zip(valores, resultados)):
            ws[f'A{i+2}'] = i + 1
            ws[f'B{i+2}'] = valor
            ws[f'C{i+2}'] = resultado[0]
            ws[f'D{i+2}'] = resultado[1] ws[f'E{i+2}'] = resultado[2]